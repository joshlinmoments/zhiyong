#!/usr/bin/env python3
"""Rebuild the Excel with the 引用 hierarchy restored.

A unit is classified 引用 when its lead line marks a 聖訓 / 經典 / 格言 quote;
otherwise 內文. Front matter (目錄 / 樞紐四訓 / 序文) sits on top as before.
First-pass auto-classification — meant for human review.
"""
import json
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = sys.argv[1]          # content.json
OUT = sys.argv[2]          # output .xlsx

# 聖訓 / 經典 / 格言 markers (checked mainly on the lead line of a unit)
CITE = re.compile(
    r"聖訓|聖諭|慈諭|"
    r"格言|"
    r"史記|漢書|後漢書|三國志|資治通鑑|"
    r"論語|孟子|大學|中庸|詩經|書經|尚書|禮記|周禮|易經|春秋|左傳|國語|戰國策|"
    r"老子|莊子|道德經|列子|韓非|淮南子|菜根譚|呂氏春秋|荀子|墨子|"
    r"子曰|孟子曰|孔子曰|古詩|古云|古聖云|古德云|古格言|憫農"
)


def kind_of(text):
    head = text.split("\n", 1)[0]
    # check lead line first; fall back to whole text for embedded classic sources
    if CITE.search(head):
        return "引用"
    # a lead-in ending in ： that names a classic anywhere on first two lines
    first2 = "\n".join(text.split("\n")[:2])
    if "：" in head and CITE.search(first2):
        return "引用"
    return "內文"


data = json.load(open(SRC, encoding="utf-8"))

wb = Workbook()
ws = wb.active
ws.title = "智勇班教材"
ws.sheet_view.showGridLines = False
PM = "PMingLiU"
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="top", wrap_text=True)
leftc = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
NCOL, LAST = 6, "F"


def merge_row(r, text, *, font=None, align=None, fill=None, height=None):
    ws.merge_cells(f"A{r}:{LAST}{r}")
    c = ws.cell(row=r, column=1, value=text)
    c.font = font or Font(name=PM, size=12)
    c.alignment = align or leftc
    if fill:
        for col in range(1, NCOL + 1):
            ws.cell(row=r, column=col).fill = fill
    if height:
        ws.row_dimensions[r].height = height


title_font = Font(name=PM, size=20, bold=True, color="1F3864")
block_font = Font(name=PM, size=13, bold=True, color="FFFFFF")
block_fill = PatternFill("solid", fgColor="305496")

r = 1
merge_row(r, data["title"], font=title_font, align=center, height=34); r += 2
merge_row(r, "目　錄", font=block_font, align=leftc, fill=block_fill); r += 1
for ch in data["chapters"]:
    merge_row(r, "　　" + ch["title"]); r += 1
r += 1
mt = data["frontMatter"]["motto"]
merge_row(r, mt["title"], font=block_font, align=leftc, fill=block_fill); r += 1
merge_row(r, "　　".join(mt["lines"]), align=center, height=22); r += 2
pf = data["frontMatter"]["preface"]
merge_row(r, pf["title"], font=block_font, align=leftc, fill=block_fill); r += 1
merge_row(r, pf["text"], align=left, height=150); r += 1
merge_row(r, pf["sign"], align=right); r += 2

hdr_fill = PatternFill("solid", fgColor="1F3864")
hdr_font = Font(name=PM, bold=True, color="FFFFFF", size=12)
headers = ["序號", "章", "節", "階層", "內容", "字數"]
table_hdr_row = r
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=r, column=c, value=h)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = border
r += 1

fill_h1 = PatternFill("solid", fgColor="FCE4D6")
fill_h2 = PatternFill("solid", fgColor="FFF2CC")
fill_q = PatternFill("solid", fgColor="E2EFDA")
from math import ceil
CPL = 37
n = 0
counts = {"章": 0, "節": 0, "內文": 0, "引用": 0}
for ch in data["chapters"]:
    rows = [("章", ch["title"], ch["title"], "")]
    for s in ch["sections"]:
        if s["title"]:
            rows.append(("節", ch["title"], s["title"], s["title"]))
        for u in s["units"]:
            lvl = kind_of(u["text"])
            rows.append((lvl, ch["title"], s["title"], u["text"]))
    for lvl, h1, h2, content in rows:
        n += 1; counts[lvl] += 1
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2, value=h1)
        ws.cell(row=r, column=3, value=h2 if lvl != "章" else "")
        ws.cell(row=r, column=4, value=lvl)
        ws.cell(row=r, column=5, value=content)
        ws.cell(row=r, column=6, value=len(content.replace("\n", "")))
        fill = {"章": fill_h1, "節": fill_h2, "引用": fill_q}.get(lvl)
        for c in range(1, NCOL + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.font = Font(name=PM, size=12, bold=(lvl in ("章", "節")))
            cell.alignment = center if c in (1, 4, 6) else left
            if fill:
                cell.fill = fill
        vlines = sum(max(1, ceil(len(seg) / CPL)) for seg in content.split("\n"))
        ws.row_dimensions[r].height = min(409, max(20, vlines * 16 + 3))
        r += 1

for col, w in {"A": 6, "B": 22, "C": 22, "D": 7, "E": 80, "F": 7}.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = f"A{table_hdr_row + 1}"
ws.auto_filter.ref = f"A{table_hdr_row}:F{r - 1}"
wb.save(OUT)
print("Saved:", OUT)
print("counts:", counts)
